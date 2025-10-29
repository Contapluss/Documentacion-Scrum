from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.models.generated import Empresa, Trabajador, DatosTrabajador, Territorial, Contrato
from app.schemas.pdf_contrato import PDFContratoRequest, PDFContratoResponse
from app.schemas.pdf_termino_contrato import PDFTerminoContratoRequest, PDFTerminoContratoResponse
from app.services.pdf_generator import PDFContratoGenerator, PDFTerminoContratoGenerator
from app.services.dependencies import get_current_user

router = APIRouter(prefix="/contrato", tags=["Contrato"])


@router.post("/generate-pdf")
def generate_contrato_pdf(
    pdf_data: PDFContratoRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] not in [1, 2]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para generar contratos"
        )

    try:
        # Obtener empresa_id del usuario autenticado
        empresa_id = current_user["empresa_id"]

        # Obtener empresa
        empresa = db.query(Empresa).filter(Empresa.id_empresa == empresa_id).first()
        if not empresa:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Empresa no encontrada"
            )

        # Construir los datos para el PDF con información de la empresa
        class PDFDataForGenerator:
            pass

        pdf_generator_data = PDFDataForGenerator()
        pdf_generator_data.ciudad_firma = pdf_data.ciudad_firma
        pdf_generator_data.fecha_contrato = pdf_data.fecha_contrato
        pdf_generator_data.empresa_nombre = empresa.nombre_fantasia
        pdf_generator_data.empresa_rut = f"{empresa.rut_empresa}-{empresa.DV_rut}"
        pdf_generator_data.representante_legal = pdf_data.representante_legal
        pdf_generator_data.rut_representante = pdf_data.rut_representante
        pdf_generator_data.domicilio_representante = pdf_data.domicilio_representante
        pdf_generator_data.nombre_trabajador = pdf_data.nombre_trabajador
        pdf_generator_data.nacionalidad_trabajador = pdf_data.nacionalidad_trabajador
        pdf_generator_data.rut_trabajador = pdf_data.rut_trabajador
        pdf_generator_data.estado_civil_trabajador = pdf_data.estado_civil_trabajador
        pdf_generator_data.fecha_nacimiento_trabajador = pdf_data.fecha_nacimiento_trabajador
        pdf_generator_data.domicilio_trabajador = pdf_data.domicilio_trabajador
        pdf_generator_data.cargo_trabajador = pdf_data.cargo_trabajador
        pdf_generator_data.lugar_trabajo = pdf_data.lugar_trabajo
        pdf_generator_data.sueldo = pdf_data.sueldo
        pdf_generator_data.jornada = pdf_data.jornada
        pdf_generator_data.descripcion_jornada = pdf_data.descripcion_jornada
        pdf_generator_data.clausulas = pdf_data.clausulas

        # Crear instancia del generador de PDF
        pdf_generator = PDFContratoGenerator()

        # Generar el PDF
        pdf_path = pdf_generator.generate_pdf(pdf_generator_data)

        # Verificar que el archivo se creó correctamente
        if not os.path.exists(pdf_path):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Error al generar el PDF"
            )

        # Devolver el archivo PDF como respuesta
        return FileResponse(
            path=pdf_path,
            media_type='application/pdf',
            filename=f"contrato_{pdf_data.rut_trabajador}.pdf"
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el PDF: {str(e)}"
        )


@router.post("/generate-pdf-termino")
def generate_termino_contrato_pdf(
    pdf_data: PDFTerminoContratoRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] not in [1, 2]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para generar cartas de término"
        )

    try:
        # Obtener empresa_id del usuario autenticado
        empresa_id = current_user["empresa_id"]

        # Validar que el RUT sea numérico
        if not pdf_data.rut_trabajador.isdigit():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El RUT debe contener solo números"
            )

        # Buscar trabajador por RUT
        trabajador_obj = None
        datos_trabajador = None

        trabajadores = db.query(Trabajador).filter(
            Trabajador.id_empresa == empresa_id
        ).all()

        for t in trabajadores:
            datos = db.query(DatosTrabajador).filter(
                DatosTrabajador.id_trabajador == t.id_trabajador,
                DatosTrabajador.rut == int(pdf_data.rut_trabajador)
            ).first()
            if datos:
                trabajador_obj = t
                datos_trabajador = datos
                break

        if not datos_trabajador:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Trabajador no encontrado"
            )

        # Obtener territorial para la comuna
        territorial = db.query(Territorial).filter(
            Territorial.id_territorial == trabajador_obj.id_territorial
        ).first()

        comuna_trabajador = territorial.comuna if territorial else "Sin comuna"

        # Obtener empresa
        empresa = db.query(Empresa).filter(Empresa.id_empresa == empresa_id).first()
        if not empresa:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Empresa no encontrada"
            )

        # Construir los datos para el PDF
        class PDFDataForGenerator:
            pass

        pdf_generator_data = PDFDataForGenerator()
        pdf_generator_data.ciudad = pdf_data.ciudad
        pdf_generator_data.fecha_carta = pdf_data.fecha_carta
        pdf_generator_data.empresa_nombre = empresa.nombre_fantasia
        pdf_generator_data.empresa_rut = f"{empresa.rut_empresa}-{empresa.DV_rut}"
        pdf_generator_data.nombre_trabajador = f"{datos_trabajador.nombre} {datos_trabajador.apellido_paterno} {datos_trabajador.apellido_materno}"
        pdf_generator_data.rut_trabajador = f"{datos_trabajador.rut}-{datos_trabajador.DV_rut}"
        pdf_generator_data.direccion_trabajador = datos_trabajador.direccion_real
        pdf_generator_data.comuna_trabajador = comuna_trabajador
        pdf_generator_data.fecha_termino = pdf_data.fecha_termino
        pdf_generator_data.articulo_causal = pdf_data.articulo_causal
        pdf_generator_data.descripcion_causal = pdf_data.descripcion_causal
        pdf_generator_data.fundamentacion = pdf_data.fundamentacion
        pdf_generator_data.lugar_pago_finiquito = pdf_data.lugar_pago_finiquito
        pdf_generator_data.telefono_notaria = pdf_data.telefono_notaria

        # Crear instancia del generador de PDF
        pdf_generator = PDFTerminoContratoGenerator()

        # Generar el PDF
        pdf_path = pdf_generator.generate_pdf(pdf_generator_data)

        # Verificar que el archivo se creó correctamente
        if not os.path.exists(pdf_path):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Error al generar el PDF"
            )

        # Devolver el archivo PDF como respuesta
        return FileResponse(
            path=pdf_path,
            media_type='application/pdf',
            filename=f"termino_contrato_{pdf_data.rut_trabajador}.pdf"
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el PDF: {str(e)}"
        )


@router.get("/generate-list-contracts")
def generate_list_contracts(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["rol"] not in [1, 2]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="No tienes permisos para generar listado de contratos"
        )

    try:
        # Obtener empresa_id del usuario autenticado
        empresa_id = current_user["empresa_id"]

        # Obtener todos los contratos de trabajadores de la empresa
        contratos = db.query(Contrato).join(
            Trabajador, Contrato.id_trabajador == Trabajador.id_trabajador
        ).filter(
            Trabajador.id_empresa == empresa_id
        ).all()

        if not contratos:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No se encontraron contratos para esta empresa"
            )

        # Crear el archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado de Contratos"

        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = [
            "ID Contrato",
            "RUT",
            "Nombre Completo",
            "Dirección Contrato",
            "Fecha Subida",
            "Fecha Inicial",
            "Fecha Término",
            "Estado"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos
        for row_idx, contrato in enumerate(contratos, start=2):
            # Obtener datos del trabajador
            datos_trabajador = db.query(DatosTrabajador).filter(
                DatosTrabajador.id_trabajador == contrato.id_trabajador
            ).first()

            if datos_trabajador:
                nombre_completo = f"{datos_trabajador.nombre} {datos_trabajador.apellido_paterno} {datos_trabajador.apellido_materno}"
                rut = f"{datos_trabajador.rut}-{datos_trabajador.DV_rut}"
            else:
                nombre_completo = "Sin datos"
                rut = "N/A"

            # Determinar estado del contrato
            hoy = datetime.now()
            if contrato.fecha_termino:
                if contrato.fecha_termino.replace(tzinfo=None) < hoy:
                    estado = "Finalizado"
                else:
                    estado = "Vigente"
            else:
                estado = "Indefinido"

            # Escribir datos
            ws.cell(row=row_idx, column=1, value=contrato.id_contrato)
            ws.cell(row=row_idx, column=2, value=rut)
            ws.cell(row=row_idx, column=3, value=nombre_completo)
            ws.cell(row=row_idx, column=4, value=contrato.direccion_contrato or "N/A")
            ws.cell(row=row_idx, column=5, value=contrato.fecha_subida.strftime('%Y-%m-%d %H:%M') if contrato.fecha_subida else "N/A")
            ws.cell(row=row_idx, column=6, value=contrato.fecha_inicial.strftime('%Y-%m-%d') if contrato.fecha_inicial else "N/A")
            ws.cell(row=row_idx, column=7, value=contrato.fecha_termino.strftime('%Y-%m-%d') if contrato.fecha_termino else "N/A")
            ws.cell(row=row_idx, column=8, value=estado)

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12

        # Guardar archivo
        excel_dir = "generated_excels"
        os.makedirs(excel_dir, exist_ok=True)
        filename = f"listado_contratos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(excel_dir, filename)
        wb.save(filepath)

        # Devolver el archivo Excel
        return FileResponse(
            path=filepath,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=filename
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar el Excel: {str(e)}"
        )
